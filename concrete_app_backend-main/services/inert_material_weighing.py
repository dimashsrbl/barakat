from utils.unitofwork import IUnitOfWork
from models.inert_material_request import InertRequestStatusEnum
from models.weighing import Weighing
from models.transport import Transport
from models.detail import Detail
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

class InertMaterialWeighingService:
    async def process_weighing(self, uow: IUnitOfWork, plate_number: str, weight: int, operator_id: int):
        async with uow:
            # Найти транспорт по номеру
            transport = await uow.transport.find_one_or_none(plate_number=plate_number)
            if not transport:
                raise Exception("Транспорт с таким номером не найден")
            
            # Получить пользователя для определения его компании
            user = await uow.user.find_one_or_none(id=operator_id)
            if not user or not user.company_id:
                raise Exception("У пользователя не указана компания!")

            # --- ДОБАВЛЕНО: ищем активную заявку от поставщика по plate_number ---
            inert_request = await uow.inert_material_request.find_one_or_none(transport_id=transport.id, status=InertRequestStatusEnum.active)
            
            # Проверить, есть ли уже отвес по этой машине (НЕ по заявке!)
            weighing = await uow.weighing.find_one_or_none(transport_id=transport.id, is_depend=False, is_finished=False)
            if not weighing:
                if inert_request:
                    # Используем detail и все id из заявки
                    detail_id = inert_request.detail_id
                    seller_company_id = user.company_id
                    client_company_id = 2  # Баракат
                    material_id = inert_request.material_id
                    print(f"[DEBUG] Создаём отвес: detail_id={detail_id}, inert_request_id={inert_request.id}, transport_id={transport.id}")
                    # Первый взвес (брутто) — создаём независимый отвес с detail_id и inert_request_id
                    weighing = await uow.weighing.create({
                        'transport_id': transport.id,
                        'detail_id': detail_id,
                        'brutto_weight': weight,  # ИЗМЕНЕНО: первый взвес теперь брутто
                        'tare_weight': None,      # ИЗМЕНЕНО: тара пока пустая
                        'first_at': datetime.utcnow(),
                        'first_operator_id': operator_id,
                        'is_depend': False,
                        'is_finished': False,
                        'is_active': True,
                        'inert_request_id': inert_request.id,
                    })
                else:
                    # Первый взвес (брутто) — создаём detail для независимого отвеса (старое поведение)
                    detail = await uow.detail.create({
                        'seller_company_id': user.company_id,
                        'client_company_id': 2,  # id компании 'Баракат'
                        'material_id': 1,  # заглушка, можно доработать
                        'construction_id': None,
                        'object_id': None,
                        'cone_draft_default': None,
                    })
                    weighing = await uow.weighing.create({
                        'transport_id': transport.id,
                        'detail_id': detail.id,
                        'brutto_weight': weight,  # ИЗМЕНЕНО: первый взвес теперь брутто
                        'tare_weight': None,      # ИЗМЕНЕНО: тара пока пустая
                        'first_at': datetime.utcnow(),
                        'first_operator_id': operator_id,
                        'is_depend': False,
                        'is_finished': False,
                        'is_active': True,
                    })
                await uow.commit()
                return {"message": "Первый взвес (брутто) записан", "weighing_id": weighing.id}
            else:
                # Второй взвес — обновляем запись, фиксируем тару
                if weighing.tare_weight:
                    raise Exception("Уже есть второй взвес для этой машины!")
                print(f"[DEBUG] Второй взвес: weighing_id={weighing.id}, inert_request_id={getattr(weighing, 'inert_request_id', None)}")
                weighing = await uow.weighing.update(weighing.id, {
                    'tare_weight': weight,  # ИЗМЕНЕНО: второй взвес теперь тара
                    'second_at': datetime.utcnow(),
                    'second_operator_id': operator_id,
                    'is_finished': True,
                    'inert_request_id': getattr(weighing, 'inert_request_id', None),
                })
                # Генерация накладной
                file_path = await self.generate_invoice(uow, weighing.id)
                # --- ДОБАВЛЕНО: если есть заявка и лабораторные поля заполнены, меняем статус заявки на finished ---
                if getattr(weighing, 'inert_request_id', None):
                    # Получаем актуальный отвес после обновления
                    updated_weighing = await uow.weighing.find_one_or_none(id=weighing.id)
                    if updated_weighing and updated_weighing.weediness is not None and updated_weighing.silo_number is not None:
                        from services.inert_material_request import InertMaterialRequestService
                        await InertMaterialRequestService().finish(uow, updated_weighing.inert_request_id)
                        await uow.commit()
                # --- КОНЕЦ ДОБАВЛЕНИЯ ---
                await uow.commit()
                return {
                    "message": "Второй взвес (тара) записан, отвес завершён", 
                    "weighing_id": weighing.id, 
                    "invoice_path": file_path
                }

    async def generate_invoice(self, uow: IUnitOfWork, weighing_id: int) -> str:
        # Получаем все нужные данные
        weighing = await uow.weighing.find_one_or_none(id=weighing_id)
        if not weighing:
            raise Exception("Отвес не найден")
        request = await uow.inert_material_request.find_one_or_none(id=weighing.inert_request_id)
        transport = await uow.transport.find_one_or_none(id=weighing.transport_id)
        
        # Проверяем, есть ли заявка и получаем carrier_id
        carrier = None
        if request and request.carrier_id:
            carrier = await uow.carrier.find_one_or_none(id=request.carrier_id)
        elif transport and transport.carrier_id:
            carrier = await uow.carrier.find_one_or_none(id=transport.carrier_id)
            
        material = await uow.material.find_one_or_none(id=request.material_id) if request else None
        operator1 = await uow.user.find_one_or_none(id=weighing.first_operator_id)
        operator2 = await uow.user.find_one_or_none(id=weighing.second_operator_id)

        netto = weighing.brutto_weight - weighing.tare_weight if weighing.brutto_weight and weighing.tare_weight else None

        wb = Workbook()
        ws = wb.active
        ws.title = "Накладная"
        ws['A1'] = "Накладная на отгрузку инертного материала"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        ws['A2'] = f"Номер заявки: {request.id if request else 'Нет заявки'}"
        ws['A3'] = f"Дата первого взвешивания: {weighing.first_at.strftime('%d.%m.%Y %H:%M')}"
        ws['A4'] = f"Дата второго взвешивания: {weighing.second_at.strftime('%d.%m.%Y %H:%M')}"
        ws['A5'] = f"Поставщик: {operator1.fullname if operator1 else ''}"
        ws['A6'] = f"Перевозчик: {carrier.name if carrier else ''}"
        ws['A7'] = f"Гос. номер: {transport.plate_number if transport else ''}"
        ws['A8'] = f"Материал: {material.name if material else ''}"
        ws['A9'] = f"Тара: {weighing.tare_weight or ''} кг"
        ws['A10'] = f"Брутто: {weighing.brutto_weight or ''} кг"
        ws['A11'] = f"Нетто: {netto or ''} кг"
        ws['A12'] = f"Оператор 1: {operator1.fullname if operator1 else ''}"
        ws['A13'] = f"Оператор 2: {operator2.fullname if operator2 else ''}"
        for i in range(1, 14):
            ws[f'A{i}'].alignment = Alignment(horizontal="left")
        # Сохраняем файл
        os.makedirs('media', exist_ok=True)
        file_path = f"media/invoice_{weighing.id}.xlsx"
        wb.save(file_path)
        return file_path 