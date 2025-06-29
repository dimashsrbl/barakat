import base64
import datetime
import io
from datetime import timedelta
from math import ceil

from aiohttp import ClientConnectionError
from openpyxl.cell import Cell, MergedCell
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.writer.excel import save_workbook
from sqlalchemy import select, func, extract
from sqlalchemy.orm import aliased

from db.db import db_helper
from exceptions import BadRequestException
from models.carrier import Carrier
from models.company import Company
from models.concrete_mixing_plant import ConcreteMixingPlant
from models.detail import Detail
from models.material import Material
from models.material_type import MaterialType
from models.object import Object
from models.transport import Transport
from models.user import User
from models.weighing import Weighing
from schemas.company import CompanyFuncEnum
from utils.cmp_integration import get_data_for_materials_report


class ReportService:
    def set_border(self, ws, cell_range, font_size: int = 11):
        thin = Side(border_style="thin", color="000000")
        fontStyle = Font(size=font_size)
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                cell.font = fontStyle

    def set_bold(self, ws, cell_range):
        for row in ws[cell_range]:
            for cell in row:
                cell.font = Font(bold=True)

    def set_print_settings(self, ws):
        cm = 2.54
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins = PageMargins(left=0.6 / cm, right=0.6 / cm, top=1, bottom=1, header=0.7 / cm, footer=0.7 / cm)

    def to_bytes(self, wb):
        bytes_io = io.BytesIO()
        save_workbook(wb, bytes_io)
        bytes_data = bytes_io.getvalue()
        return base64.b64encode(bytes_data).decode()

    async def get_summary(self, from_date, to_date, seller_companies, client_companies, **filter_by) \
            -> str:
        async with db_helper.get_db_session() as session:
            seller_companies_list, client_companies_list = [], []

            if seller_companies:
                try:
                    seller_companies_list = [int(id) for id in seller_companies.split(",")]
                except:
                    raise BadRequestException("seller companies are incorrect")
            if client_companies:
                try:
                    client_companies_list = [int(id) for id in client_companies.split(",")]
                except:
                    raise BadRequestException("client companies are incorrect")

            seller = aliased(Company)
            client = aliased(Company)

            stmt = select(Company.id).filter_by(company_func=CompanyFuncEnum.our)
            raw = await session.execute(stmt)
            our_company_ids = raw.scalars().all()

            if len(client_companies_list) and all(
                    company in our_company_ids for company in client_companies_list) and not any(
                company in our_company_ids for company in seller_companies_list):
                # group by SELLER and material
                stmt = (
                    select(
                        seller.company_type.label('company_type'),
                        seller.name.label('company'),
                        Carrier.name.label('carrier'),
                        Material.name.label('material'),
                        Weighing.is_return.label('is_return'),
                        func.sum(Weighing.clean_weight).label('clean'),
                    ).
                    join(
                        Detail, Weighing.detail_id == Detail.id
                    ).
                    join(
                        seller, Detail.seller_company_id == seller.id
                    ).
                    join(
                        client, Detail.client_company_id == client.id
                    ).
                    join(
                        Material, Detail.material_id == Material.id
                    ).
                    join(
                        Transport, Weighing.transport_id == Transport.id
                    ).
                    outerjoin(
                        Carrier, Transport.carrier_id == Carrier.id
                    ).
                    filter(
                        Weighing.is_active == True
                        and
                        Weighing.is_finished == True,
                        Weighing.is_depend == False,
                        (not seller_companies_list or seller.id.in_(seller_companies_list)),
                        (not client_companies_list or client.id.in_(client_companies_list)),
                        Weighing.clean_weight.isnot(None),
                        Weighing.second_at.between(from_date, to_date),
                    ).
                    group_by(
                        seller.company_type, seller.name, Carrier.name, Material.name, Weighing.is_return,
                    )
                )
            elif len(seller_companies_list) and all(
                    company in our_company_ids for company in seller_companies_list) and not any(
                company in our_company_ids for company in client_companies_list):
                # group by CLIENT and material
                stmt = (
                    select(
                        client.company_type.label('company_type'),
                        client.name.label('company'),
                        Carrier.name.label('carrier'),
                        Material.name.label('material'),
                        Weighing.is_return.label('is_return'),
                        func.sum(Weighing.clean_weight).label('clean'),
                    ).
                    join(
                        Detail, Weighing.detail_id == Detail.id
                    ).
                    join(
                        seller, Detail.seller_company_id == seller.id
                    ).
                    join(
                        client, Detail.client_company_id == client.id
                    ).
                    join(
                        Material, Detail.material_id == Material.id
                    ).
                    join(
                        Transport, Weighing.transport_id == Transport.id
                    ).
                    outerjoin(
                        Carrier, Transport.carrier_id == Carrier.id
                    ).
                    filter(
                        Weighing.is_active == True
                        and
                        Weighing.is_finished == True,
                        (not seller_companies_list or seller.id.in_(seller_companies_list)),
                        (not client_companies_list or client.id.in_(client_companies_list)),
                        Weighing.clean_weight.isnot(None),
                        Weighing.second_at.between(from_date, to_date),
                    ).
                    group_by(
                        client.company_type, client.name, Carrier.name, Material.name, Weighing.is_return
                    )
                )
            else:
                raise BadRequestException("Неверные данные в сводном отчете")

            raw = await session.execute(stmt)
            results = raw.mappings().all()

            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Data"
            ws1.merge_cells('A1:D1')

            width = 85
            ws1.column_dimensions['A'].width = width / 4 * 3
            ws1.column_dimensions['b'].width = width / 4
            ws1.column_dimensions['c'].width = width / 2

            from_date_title = from_date.strftime("%d-%m-%Y %H:%M")
            to_date_title = to_date.strftime("%d-%m-%Y %H:%M")

            ws1['A1'] = f"Сводный отчет по отвесам ТОО 'BARAKAT INVEST' с {from_date_title} по {to_date_title}"

            ws1['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws1['A2'] = f"Контрагент"
            ws1['A2'].alignment = Alignment(horizontal="center", vertical="center")
            ws1['B2'] = f"Перевозчик"
            ws1['B2'].alignment = Alignment(horizontal="center", vertical="center")
            ws1['C2'] = f"Материал"
            ws1['C2'].alignment = Alignment(horizontal="center", vertical="center")
            ws1['D2'] = f"Вес"
            ws1['D2'].alignment = Alignment(horizontal="center", vertical="center")

            counter = 3
            clean_sum = 0
            last_cell: str

            for result in results:
                ws1[f'A{counter}'] = f"{result['company_type'].value} {result['company']}"
                ws1[f'B{counter}'] = result['carrier']
                if result['is_return']:
                    ws1[f'C{counter}'] = f"{result['material']} - Возврат"
                else:
                    ws1[f'C{counter}'] = result['material']
                ws1[f'D{counter}'] = result['clean']

                clean_sum += result['clean']
                counter += 1

            ws1.merge_cells(f'A{counter}:C{counter}')
            ws1[f'D{counter}'] = clean_sum

            self.set_border(ws1, f'A1:D{counter}')
            self.set_print_settings(ws1)

            ws1[f'A{counter}'] = "Итого"
            ws1[f'A{counter}'].alignment = Alignment(horizontal="right", vertical="center", wrapText=True)
            await session.close()
            return self.to_bytes(wb)

    async def get_detail(self, from_date, to_date, materials, seller_companies, client_companies, carriers, **filter_by) \
            -> str:
        async with db_helper.get_db_session() as session:
            materials_list, seller_companies_list, client_companies_list, carriers_list = [], [], [], []

            if materials:
                try:
                    materials_list = [int(id) for id in materials.split(",")]
                except:
                    raise BadRequestException("material are incorrect")
            if seller_companies:
                try:
                    seller_companies_list = [int(id) for id in seller_companies.split(",")]
                except:
                    raise BadRequestException("seller companies are incorrect")
            if client_companies:
                try:
                    client_companies_list = [int(id) for id in client_companies.split(",")]
                except:
                    raise BadRequestException("client companies are incorrect")
            if carriers:
                try:
                    carriers_list = [int(id) for id in carriers.split(",")]
                except:
                    raise BadRequestException("carriers are incorrect")

            seller = aliased(Company)
            client = aliased(Company)

            stmt = (
                select(
                    Weighing.first_at,
                    Weighing.second_at,
                    # Поставщик
                    seller.company_type.label("seller_company_type"),
                    seller.name.label("seller_name"),
                    # Заказчик
                    client.company_func.label("client_company_func"),
                    client.company_type.label("client_company_type"),
                    client.name.label("client_name"),
                    Carrier.name.label("carrier_name"),
                    Transport.plate_number,
                    Material.name.label("material_name"),
                    Weighing.silo_number,
                    Weighing.clean_weight,
                    Weighing.brutto_weight,
                    Weighing.tare_weight,
                    Weighing.return_note,
                    User.fullname
                    # 15
                ).
                join(
                    Detail, Weighing.detail_id == Detail.id
                ).
                join(
                    seller, Detail.seller_company_id == seller.id
                ).
                join(
                    client, Detail.client_company_id == client.id
                ).
                join(
                    Transport, Weighing.transport_id == Transport.id
                ).
                outerjoin(
                    Carrier, Transport.carrier_id == Carrier.id
                ).
                join(
                    Material, Detail.material_id == Material.id
                ).
                join(
                    User, Weighing.second_operator_id == User.id
                ).
                filter(
                    Weighing.is_active == True,
                    Weighing.is_finished == True,
                    (not materials_list or Material.id.in_(materials_list)),
                    (not seller_companies_list or seller.id.in_(seller_companies_list)),
                    (not client_companies_list or client.id.in_(client_companies_list)),
                    (not carriers_list or Carrier.id.in_(carriers_list)),
                    Weighing.clean_weight.isnot(None),
                    Weighing.second_at.between(from_date, to_date)
                ).
                order_by(
                    Weighing.second_at
                )
            )

            raw = await session.execute(stmt)
            results = raw.mappings().all()

            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Data"
            ws1.merge_cells('A1:L1')

            width = 100
            height = 100
            ws1.column_dimensions['A'].width = width / 10 / 1.15
            ws1.row_dimensions[1].height = height / 2

            # Дата и время заезда/выезда
            ws1.column_dimensions['B'].width = width / 6 / 1.15

            # Контрагент
            ws1.column_dimensions['C'].width = width / 6 / 1.15

            # Перевозчик
            ws1.column_dimensions['D'].width = width / 6 / 1.15

            # Номер машины
            ws1.column_dimensions['E'].width = width / 8 / 1.15

            # Груз
            ws1.column_dimensions['F'].width = width / 5 / 1.15

            # Нетто
            ws1.column_dimensions['G'].width = width / 10 / 1.15

            # Брутто
            ws1.column_dimensions['H'].width = width / 10 / 1.15

            # Тара
            ws1.column_dimensions['I'].width = width / 10 / 1.15

            # Оператор
            ws1.column_dimensions['J'].width = width / 8

            # Примечание
            ws1.column_dimensions['K'].width = width / 6

            # Примечание
            ws1.column_dimensions['L'].width = width / 14

            from_date_title = from_date.strftime("%d-%m-%Y %H:%M")
            to_date_title = to_date.strftime("%d-%m-%Y %H:%M")

            ws1['A1'] = (f"Детальный отчет по отвесам ТОО\n"
                         f"'BARAKAT INVEST' с {from_date_title} по {to_date_title}")

            ws1['A1'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

            ws1['A2'] = f"№"
            ws1['A2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['B2'] = (f"Дата и время\n"
                         f"заезда/выезда")
            ws1['B2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['C2'] = f"Контрагент"
            ws1['C2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['D2'] = f"Перевозчик"
            ws1['D2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['E2'] = (f"Номер\n"
                         f"машины")
            ws1['E2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['F2'] = f"Груз"
            ws1['F2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['G2'] = f"Нетто"
            ws1['G2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['H2'] = f"Брутто"
            ws1['H2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['I2'] = f"Тара"
            ws1['I2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['J2'] = f"Оператор"
            ws1['J2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['K2'] = f"Примечание"
            ws1['K2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['L2'] = f"Номер силоса"
            ws1['L2'].alignment = Alignment(horizontal="center", vertical="center")

            counter = 3
            clean_sum = 0
            last_cell: str

            for result in results:
                from_date_data = result['first_at'].strftime("%d-%m-%Y %H:%M")
                to_date_data = result['second_at'].strftime("%d-%m-%Y %H:%M")

                # Счетчик
                ws1[f'A{counter}'] = counter - 2

                # Дата от и до
                ws1[f'B{counter}'] = (f"{from_date_data}\n"
                                      f"{to_date_data}")

                # Функция заказчика
                if result['client_company_func'] == CompanyFuncEnum.our:
                    ws1[f'C{counter}'] = f"{result['seller_company_type'].value} {result['seller_name']}"
                else:
                    ws1[f'C{counter}'] = f"{result['client_company_type'].value} {result['client_name']}"
                # Перевозчик
                ws1[f'D{counter}'] = result['carrier_name']

                # Номер машины
                ws1[f'E{counter}'] = result['plate_number']

                # Груз
                ws1[f'F{counter}'] = result['material_name']

                # Нетто
                ws1[f'G{counter}'] = result['clean_weight']
                ws1[f'G{counter}'].font = Font(bold=False)

                # Подсчет тотала
                clean_sum += result['clean_weight']

                # Брутто
                ws1[f'H{counter}'] = result['brutto_weight']

                # Тара
                ws1[f'I{counter}'] = result['tare_weight']

                # Оператор
                ws1[f'J{counter}'] = result['fullname']

                # Примечание
                ws1[f'K{counter}'] = result['return_note']

                # Номер силоса
                ws1[f'L{counter}'] = result['silo_number']

                counter += 1

            # Тотал
            ws1[f'F{counter}'] = "Итого"
            ws1[f'F{counter}'].font = Font(bold=True)
            ws1[f'G{counter}'] = clean_sum
            ws1[f'G{counter}'].font = Font(bold=True)

            self.set_border(ws1, f'A1:L{counter - 1}', 8)
            self.set_print_settings(ws1)
            await session.close()
            return self.to_bytes(wb)

    async def get_detail_by_deleted_or_adjusted(self, from_date, to_date) -> str:
        async with db_helper.get_db_session() as session:
            seller = aliased(Company)
            client = aliased(Company)

            stmt = (
                select(
                    Weighing.id,
                    Weighing.first_at,
                    Weighing.second_at,
                    # Поставщик
                    seller.company_type.label("seller_company_type"),
                    seller.name.label("seller_name"),
                    # Заказчик
                    client.company_func.label("client_company_func"),
                    client.company_type.label("client_company_type"),
                    client.name.label("client_name"),
                    Carrier.name.label("carrier_name"),
                    Transport.plate_number,
                    Material.name.label("material_name"),
                    Weighing.clean_weight,
                    Weighing.brutto_weight,
                    Weighing.old_brutto_weight,
                    Weighing.tare_weight,
                    Weighing.old_tare_weight,
                    Weighing.bag_details,
                    Weighing.old_bag_details,
                    Weighing.deactivate_note,
                    Weighing.adjust_note,
                    User.fullname
                ).
                join(
                    Detail, Weighing.detail_id == Detail.id
                ).
                join(
                    seller, Detail.seller_company_id == seller.id
                ).
                join(
                    client, Detail.client_company_id == client.id
                ).
                join(
                    Transport, Weighing.transport_id == Transport.id
                ).
                outerjoin(
                    Carrier, Transport.carrier_id == Carrier.id
                ).
                join(
                    Material, Detail.material_id == Material.id
                ).
                outerjoin(
                    User, Weighing.second_operator_id == User.id
                ).
                filter(
                    # Weighing.is_finished == True,
                    Weighing.is_depend == False,
                    # Weighing.clean_weight.isnot(None),
                    Weighing.second_at.between(from_date, to_date),
                ).
                order_by(
                    Weighing.second_at
                )
            )

            deactivated_stmt = stmt.filter(Weighing.is_active == False)
            raw = await session.execute(deactivated_stmt)
            deactivated_results = raw.mappings().all()

            adjusted_stmt = stmt.filter(Weighing.is_adjusted == True)
            raw = await session.execute(adjusted_stmt)
            adjusted_results = raw.mappings().all()

            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Data"

            width = 100
            height = 100
            ws1.column_dimensions['A'].width = width / 10 / 1.15
            ws1.row_dimensions[1].height = height / 2

            # Дата и время заезда/выезда
            ws1.column_dimensions['B'].width = width / 6 / 1.15

            # Контрагент
            ws1.column_dimensions['C'].width = width / 6 / 1.15

            # Перевозчик
            ws1.column_dimensions['D'].width = width / 6 / 1.15

            # Номер машины
            ws1.column_dimensions['E'].width = width / 8 / 1.15

            # Груз
            ws1.column_dimensions['F'].width = width / 5 / 1.15

            # Нетто
            ws1.column_dimensions['G'].width = width / 10 / 1.15

            # Брутто
            ws1.column_dimensions['H'].width = width / 10 / 1.15

            # Тара
            ws1.column_dimensions['I'].width = width / 10 / 1.15

            # Оператор
            ws1.column_dimensions['J'].width = width / 6 / 1.15

            # Примечание/Оператор
            ws1.column_dimensions['K'].width = width / 4

            # Примечание
            ws1.column_dimensions['L'].width = width / 4

            # Старое брутто
            ws1.column_dimensions['M'].width = width / 10

            # Старая тара
            ws1.column_dimensions['N'].width = width / 10

            # Старое кол-во мешков
            ws1.column_dimensions['O'].width = width / 6

            from_date_title = from_date.strftime("%d-%m-%Y %H:%M")
            to_date_title = to_date.strftime("%d-%m-%Y %H:%M")

            ws1.merge_cells('A1:K1')
            ws1['A1'] = (f"Отчет по удаленным независимым отвесам ТОО\n"
                         f"'BARAKAT INVEST' с {from_date_title} по {to_date_title}")
            ws1['A1'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

            ws1['A2'] = f"№"
            ws1['A2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['B2'] = (f"Дата и время\n"
                         f"заезда/выезда")
            ws1['B2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['C2'] = f"Контрагент"
            ws1['C2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['D2'] = f"Перевозчик"
            ws1['D2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['E2'] = (f"Номер\n"
                         f"машины")
            ws1['E2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['F2'] = f"Груз"
            ws1['F2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['G2'] = f"Нетто"
            ws1['G2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['H2'] = f"Брутто"
            ws1['H2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['I2'] = f"Тара"
            ws1['I2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['J2'] = f"Оператор"
            ws1['J2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['K2'] = f"Примечание"
            ws1['K2'].alignment = Alignment(horizontal="center", vertical="center")

            counter = 3
            clean_sum = 0
            last_cell: str

            for result in deactivated_results:
                from_date_data = result['first_at'].strftime("%d-%m-%Y %H:%M")
                to_date_data = result['second_at'].strftime("%d-%m-%Y %H:%M")

                # Счетчик
                ws1[f'A{counter}'] = counter - 2

                # Дата от и до
                ws1[f'B{counter}'] = (f"{from_date_data}\n"
                                      f"{to_date_data}")

                # Функция заказчика
                if result['client_company_func'] == CompanyFuncEnum.our:
                    ws1[f'C{counter}'] = f"{result['seller_company_type'].value} {result['seller_name']}"
                else:
                    ws1[f'C{counter}'] = f"{result['client_company_type'].value} {result['client_name']}"
                # Перевозчик
                ws1[f'D{counter}'] = result['carrier_name']

                # Номер машины
                ws1[f'E{counter}'] = result['plate_number']

                # Груз
                ws1[f'F{counter}'] = result['material_name']

                # Нетто
                ws1[f'G{counter}'] = result['clean_weight']
                ws1[f'G{counter}'].font = Font(bold=False)

                # Подсчет тотала
                clean_sum += result['clean_weight'] or 0

                # Брутто
                ws1[f'H{counter}'] = result['brutto_weight']

                # Тара
                ws1[f'I{counter}'] = result['tare_weight']

                # Оператор
                ws1[f'J{counter}'] = result['fullname']

                # Примечание
                ws1[f'K{counter}'] = result['deactivate_note']

                counter += 1

            # Тотал
            ws1[f'F{counter}'] = "Итого"
            ws1[f'F{counter}'].font = Font(bold=True)
            ws1[f'G{counter}'] = clean_sum
            ws1[f'G{counter}'].font = Font(bold=True)

            clean_sum = 0
            self.set_border(ws1, f'A1:K{counter - 1}', 8)
            counter += 2
            start_row = counter

            ws1.merge_cells(f'A{counter}:O{counter}')
            ws1[f'A{counter}'] = (f"Отчет по корректированным независимым отвесам ТОО\n"
                                  f"'BARAKAT INVEST' с {from_date_title} по {to_date_title}")
            ws1.row_dimensions[counter].height = height / 2

            ws1[f'A{counter}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            counter += 1
            ws1[f'A{counter}'] = f"№"
            ws1[f'A{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'B{counter}'] = (f"Дата и время\n"
                                  f"заезда/выезда")
            ws1[f'B{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'C{counter}'] = f"Контрагент"
            ws1[f'C{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'D{counter}'] = f"Перевозчик"
            ws1[f'D{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'E{counter}'] = (f"Номер\n"
                                  f"машины")
            ws1[f'E{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'F{counter}'] = f"Груз"
            ws1[f'F{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'G{counter}'] = f"Нетто"
            ws1[f'G{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'H{counter}'] = f"Брутто"
            ws1[f'H{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'I{counter}'] = f"Тара"
            ws1[f'I{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'J{counter}'] = f"Кол-во мешков"
            ws1[f'J{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'K{counter}'] = f"Оператор"
            ws1[f'K{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'L{counter}'] = f"Примечание"
            ws1[f'L{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'M{counter}'] = f"Старое брутто"
            ws1[f'M{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'N{counter}'] = f"Старая тара"
            ws1[f'N{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'O{counter}'] = f"Старое кол-во мешков"
            ws1[f'O{counter}'].alignment = Alignment(horizontal="center", vertical="center")
            counter += 1

            for result in adjusted_results:
                from_date_data = result['first_at'].strftime("%d-%m-%Y %H:%M")
                to_date_data = result['second_at'].strftime("%d-%m-%Y %H:%M")

                # Счетчик
                ws1[f'A{counter}'] = counter - 6

                # Дата от и до
                ws1[f'B{counter}'] = (f"{from_date_data}\n"
                                      f"{to_date_data}")

                # Функция заказчика
                if result['client_company_func'] == CompanyFuncEnum.our:
                    ws1[f'C{counter}'] = f"{result['seller_company_type'].value} {result['seller_name']}"
                else:
                    ws1[f'C{counter}'] = f"{result['client_company_type'].value} {result['client_name']}"

                # Перевозчик
                ws1[f'D{counter}'] = result['carrier_name']

                # Номер машины
                ws1[f'E{counter}'] = result['plate_number']

                # Груз
                ws1[f'F{counter}'] = result['material_name']

                # Нетто
                ws1[f'G{counter}'] = result['clean_weight']
                ws1[f'G{counter}'].font = Font(bold=False)

                # Подсчет тотала
                clean_sum += result['clean_weight']

                # Брутто
                ws1[f'H{counter}'] = result['brutto_weight']

                # Тара
                ws1[f'I{counter}'] = result['tare_weight']

                # Кол-во мешков
                ws1[f'J{counter}'] = result['bag_details']

                # Оператор
                ws1[f'K{counter}'] = result['fullname']

                # Примечание
                ws1[f'L{counter}'] = result['adjust_note']

                # Старое брутто
                ws1[f'M{counter}'] = result['old_brutto_weight']

                # Старое тара
                ws1[f'N{counter}'] = result['old_brutto_weight']

                # Старое кол-во мешков
                ws1[f'O{counter}'] = result['old_bag_details']

                counter += 1

            # Тотал
            ws1[f'F{counter}'] = "Итого"
            ws1[f'F{counter}'].font = Font(bold=True)
            ws1[f'G{counter}'] = clean_sum
            ws1[f'G{counter}'].font = Font(bold=True)

            self.set_border(ws1, f'A{start_row}:O{counter - 1}', 8)
            self.set_print_settings(ws1)
            await session.close()
            return self.to_bytes(wb)

    async def get_independent_by_materials(self, from_date, to_date, materials) -> str:
        async with db_helper.get_db_session() as session:
            materials_list = []

            if materials:
                try:
                    materials_list = [int(id) for id in materials.split(",")]
                except:
                    raise BadRequestException("material are incorrect")
            else:
                stmt = (
                    select(
                        Material.id
                    )
                    .join(
                        MaterialType, Material.material_type_id == MaterialType.id
                    )
                    .filter(
                        MaterialType.is_for_independent == True,
                    )
                )
                raw = await session.execute(stmt)
                materials_list = raw.scalars().all()

            stmt = (
                select(
                    Material.name,
                    func.sum(Weighing.clean_weight).label('clean'),
                ).
                join(
                    Detail, Weighing.detail_id == Detail.id
                ).
                join(
                    Material, Detail.material_id == Material.id
                ).
                filter(
                    Weighing.is_active == True,
                    Weighing.is_finished == True,
                    Weighing.is_depend == False,
                    (not materials_list or Material.id.in_(materials_list)),
                    Weighing.clean_weight.isnot(None),
                    Weighing.second_at.between(from_date, to_date),
                ).
                group_by(
                    Material.name
                )
            )

            raw = await session.execute(stmt)
            results = raw.mappings().all()

            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Data"
            ws1.merge_cells('A1:B1')

            width = 100
            ws1.column_dimensions['A'].width = width / 2
            ws1.row_dimensions[1].height = 25

            # Дата и время заезда/выезда
            ws1.column_dimensions['B'].width = width / 2

            from_date_title = from_date.strftime("%d-%m-%Y %H:%M")
            to_date_title = to_date.strftime("%d-%m-%Y %H:%M")

            ws1['A1'] = (f"Отчет по материалам по независимым отвесам ТОО\n"
                         f"'BARAKAT INVEST' с {from_date_title} по {to_date_title}")

            ws1['A1'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

            ws1['A2'] = f"Материал"
            ws1['A2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['B2'] = f"Нетто"
            ws1['B2'].alignment = Alignment(horizontal="center", vertical="center")

            counter = 3
            clean_sum = 0
            last_cell: str

            for result in results:
                # Счетчик
                ws1[f'A{counter}'] = result['name']

                # Дата от и до
                ws1[f'B{counter}'] = result['clean']
                clean_sum += result['clean']

                counter += 1

            # Тотал
            ws1[f'A{counter}'] = "Итого"
            ws1[f'A{counter}'].font = Font(bold=True)
            ws1[f'B{counter}'] = clean_sum
            ws1[f'B{counter}'].font = Font(bold=True)

            self.set_border(ws1, f'A1:B{counter - 1}')
            self.set_print_settings(ws1)
            await session.close()
            return self.to_bytes(wb)

    async def get_dependent_summary(self, from_date, to_date, material_types, client_companies, objects, **filter_by) \
            -> str:
        async with db_helper.get_db_session() as session:
            client_companies_list, objects_list = [], []

            stmt = select(MaterialType.id).filter_by(is_for_dependent=True)
            raw = await session.execute(stmt)
            material_types_list = raw.scalars().all()

            if material_types:
                try:
                    material_types_list = [int(id) for id in material_types.split(",")]
                except:
                    raise BadRequestException("material types are incorrect")
            if client_companies:
                try:
                    client_companies_list = [int(id) for id in client_companies.split(",")]
                except:
                    raise BadRequestException("client companies are incorrect")
            if objects:
                try:
                    objects_list = [int(id) for id in objects.split(",")]
                except:
                    raise BadRequestException("objects are incorrect")

            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Data"

            title_row = 'A'
            title_column = 1
            width = 85
            ws1.column_dimensions['A'].width = width / 4 * 3
            ws1.column_dimensions['b'].width = width / 4
            ws1.column_dimensions['c'].width = width / 2
            from_date_title = from_date.strftime("%d-%m-%Y %H:%M")
            to_date_title = to_date.strftime("%d-%m-%Y %H:%M")
            counter = 3
            clean_sum = 0
            total = 0
            last_cell: str

            cells_to_merge = f'A{title_column}:D{title_column}'
            ws1.merge_cells(cells_to_merge)
            ws1[
                f"{title_row}{title_column}"] = f"Отгрузка по типам материалов ТОО 'Торговый Дом Баракат Инвест' с {from_date_title} по {to_date_title}"
            ws1[f'{title_row}{title_column}'].alignment = Alignment(horizontal="center", vertical="center")

            start_cell = f'A{title_column}'
            counter = title_column + 1

            ws1[f'A{counter}'] = f"Заказчик"
            ws1[f'A{counter}'].alignment = Alignment(horizontal="center", vertical="center")
            ws1[f'B{counter}'] = f"Объект"
            ws1[f'B{counter}'].alignment = Alignment(horizontal="center", vertical="center")
            ws1[f'C{counter}'] = f"Материал"
            ws1[f'C{counter}'].alignment = Alignment(horizontal="center", vertical="center")
            ws1[f'D{counter}'] = f"Кубатура"
            ws1[f'D{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            for material_type_id in material_types_list:
                stmt = select(MaterialType.name).filter_by(id=material_type_id)
                raw = await session.execute(stmt)
                material_type_name = raw.scalar_one_or_none()

                stmt = (
                    select(
                        Company.company_type.label('company_type'),
                        Company.name.label('company'),
                        Object.name.label('object'),
                        Material.name.label('material'),
                        func.sum(Weighing.cubature).label('cubature'),
                    ).
                    join(
                        Detail, Weighing.detail_id == Detail.id
                    ).
                    join(
                        Company, Detail.client_company_id == Company.id
                    ).
                    join(
                        Material, Detail.material_id == Material.id
                    ).
                    join(
                        Object, Detail.object_id == Object.id
                    ).
                    filter(
                        Weighing.is_active == True
                        and
                        Weighing.is_finished == True,
                        Weighing.is_depend == True,
                        (not client_companies_list or Company.id.in_(client_companies_list)),
                        (not objects_list or Object.id.in_(objects_list)),
                        Weighing.clean_weight.isnot(None),
                        Weighing.second_at.between(from_date, to_date),
                        Material.material_type_id == material_type_id,
                    ).
                    group_by(
                        Company.name, Company.company_type, Object.name, Material.name,
                    )
                )

                raw = await session.execute(stmt)
                results = raw.mappings().all()

                if not results:
                    continue
                # cells_to_merge = f'A{title_column}:D{title_column}'
                # ws1.merge_cells(cells_to_merge)
                # ws1[f"{title_row}{title_column}"] = f"Отгрузка по материалу "{material_type_name}" ТОО "Торговый Дом Баракат Инвест" с {from_date_title} по {to_date_title}"
                # ws1[f'{title_row}{title_column}'].alignment = Alignment(horizontal="center", vertical="center")
                #
                # start_cell = f'A{title_column}'
                # counter = title_column+1
                #
                # ws1[f'A{counter}'] = f"Заказчик"
                # ws1[f'A{counter}'].alignment = Alignment(horizontal="center", vertical="center")
                # ws1[f'B{counter}'] = f"Объект"
                # ws1[f'B{counter}'].alignment = Alignment(horizontal="center", vertical="center")
                # ws1[f'C{counter}'] = f"Материал"
                # ws1[f'C{counter}'].alignment = Alignment(horizontal="center", vertical="center")
                # ws1[f'D{counter}'] = f"Кубатура"
                # ws1[f'D{counter}'].alignment = Alignment(horizontal="center", vertical="center")

                for result in results:
                    counter += 1
                    ws1[f'A{counter}'] = f"{result['company_type'].value} {result['company']}"
                    ws1[f'B{counter}'] = result['object']
                    ws1[f'C{counter}'] = result['material']
                    ws1[f'D{counter}'] = result['cubature']

                    clean_sum += result['cubature']

                counter += 1
                ws1.merge_cells(f'A{counter}:C{counter}')
                last_cell = f'D{counter}'
                ws1[last_cell] = clean_sum
                self.set_border(ws1, f'{start_cell}:{last_cell}')

                ws1[f'A{counter}'] = f"Итого по типу материала {material_type_name}"
                ws1[f'A{counter}'].alignment = Alignment(horizontal="right", vertical="center", wrapText=True)

                title_column = counter + 1
                start_cell = f'A{title_column}'

                total += clean_sum
                clean_sum = 0

            ws1[f'C{counter + 2}'] = "Итого по всем типам материала"
            ws1[f'D{counter + 2}'] = total

            self.set_print_settings(ws1)
            await session.close()
            return self.to_bytes(wb)

    async def get_dependent_detail(self, from_date, to_date, materials, client_companies, objects, **filter_by) \
            -> str:
        async with db_helper.get_db_session() as session:

            materials_list, client_companies_list, objects_list = [], [], []

            if materials:
                try:
                    materials_list = [int(id) for id in materials.split(",")]
                except:
                    raise BadRequestException("material are incorrect")
            if client_companies:
                try:
                    client_companies_list = [int(id) for id in client_companies.split(",")]
                except:
                    raise BadRequestException("client companies are incorrect")
            if objects:
                try:
                    objects_list = [int(id) for id in objects.split(",")]
                except:
                    raise BadRequestException("objects are incorrect")

            stmt = (
                select(
                    Weighing.first_at,
                    Weighing.second_at,
                    Company.company_type,
                    Company.name.label('company_name'),
                    Object.name.label('object_name'),
                    Transport.plate_number,
                    Material.name.label('material_name'),
                    Carrier.name.label('carrier_name'),
                    Weighing.clean_weight,
                    Weighing.brutto_weight,
                    Weighing.tare_weight,
                    Weighing.cubature,
                    User.fullname
                ).
                join(
                    Detail, Weighing.detail_id == Detail.id
                ).
                join(
                    Company, Detail.client_company_id == Company.id
                ).
                join(
                    Object, Detail.object_id == Object.id
                ).
                join(
                    Transport, Weighing.transport_id == Transport.id
                ).
                outerjoin(
                    Carrier, Transport.carrier_id == Carrier.id
                ).
                join(
                    Material, Detail.material_id == Material.id
                ).
                join(
                    User, Weighing.second_operator_id == User.id
                ).
                filter(
                    Weighing.is_active == True,
                    Weighing.is_finished == True,
                    Weighing.is_depend == True,
                    (not materials_list or Material.id.in_(materials_list)),
                    (not client_companies_list or Company.id.in_(client_companies_list)),
                    (not objects_list or Object.id.in_(objects_list)),
                    Weighing.clean_weight.isnot(None),
                    Weighing.second_at.between(from_date, to_date)
                ).
                order_by(
                    Weighing.second_at
                )
            )

            raw = await session.execute(stmt)
            results = raw.mappings().all()

            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Data"
            ws1.merge_cells('A1:M1')

            width = 100
            height = 100
            ws1.column_dimensions['A'].width = 5
            ws1.row_dimensions[1].height = 50

            # Дата и время заезда/выезда
            ws1.column_dimensions['B'].width = width / 6

            # Заказчик
            ws1.column_dimensions['C'].width = width / 6

            # Объект
            ws1.column_dimensions['D'].width = width / 8

            # Номер машины
            ws1.column_dimensions['E'].width = width / 8.5

            # Груз
            ws1.column_dimensions['F'].width = width / 8

            # Перевозчик
            ws1.column_dimensions['G'].width = 11.7

            # Нетто
            ws1.column_dimensions['H'].width = width / 12

            # Брутто
            ws1.column_dimensions['I'].width = width / 13

            # Тара
            ws1.column_dimensions['J'].width = width / 13

            # Кубатура
            ws1.column_dimensions['K'].width = width / 11

            # Плотность
            ws1.column_dimensions['L'].width = 10.5

            # Диспетчер
            ws1.column_dimensions['M'].width = width / 8

            from_date_title = from_date.strftime("%d-%m-%Y %H:%M")
            to_date_title = to_date.strftime("%d-%m-%Y %H:%M")

            ws1[
                'A1'] = f"Детальный отчет по заявкам ТОО 'Торговый Дом Баракат Инвест' с {from_date_title} по {to_date_title}"
            ws1['A1'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

            ws1['A2'] = f"№"
            ws1['A2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['B2'] = (f"Дата и время\n"
                         f"заезда/выезда")
            ws1['B2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['C2'] = f"Заказчик"
            ws1['C2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['D2'] = f"Объект"
            ws1['D2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['E2'] = (f"Номер\n"
                         f"машины")
            ws1['E2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['F2'] = f"Марка б/р"
            ws1['F2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['G2'] = f"Перевозчик"
            ws1['G2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['H2'] = f"Нетто"
            ws1['H2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['I2'] = f"Брутто"
            ws1['I2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['J2'] = f"Тара"
            ws1['J2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['K2'] = f"Кубатура"
            ws1['K2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['L2'] = f"Плотность"
            ws1['L2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['M2'] = f"Диспетчер"
            ws1['M2'].alignment = Alignment(horizontal="center", vertical="center")
            counter = 3
            cubature_sum = 0
            last_cell: str

            for result in results:
                from_date_data = result['first_at'].strftime("%d-%m-%Y %H:%M")
                to_date_data = result['second_at'].strftime("%d-%m-%Y %H:%M")

                # Счетчик
                ws1[f'A{counter}'] = counter - 2

                # Дата от и до
                ws1[f'B{counter}'] = (f"{from_date_data}\n"
                                      f"{to_date_data}")

                # Заказчик
                ws1[f'C{counter}'] = f"{result['company_type'].value} {result['company_name']}"

                # Объект
                ws1[f'D{counter}'] = result['object_name']

                # Номер машины
                ws1[f'E{counter}'] = result['plate_number']

                # Груз
                ws1[f'F{counter}'] = result['material_name']

                # Перевозчик
                ws1[f'G{counter}'] = result['carrier_name']

                # Нетто
                ws1[f'H{counter}'] = result['clean_weight']
                # ws1[f'H{counter}'].font = Font(bold=False)

                # Брутто
                ws1[f'I{counter}'] = result['brutto_weight']

                # Тара
                ws1[f'J{counter}'] = result['tare_weight']

                # Кубатура
                ws1[f'K{counter}'] = result['cubature']

                # Плотность
                ws1[f'L{counter}'] = ceil(result['clean_weight'] / result['cubature'])

                # Оператор
                ws1[f'M{counter}'] = result['fullname']

                # Подсчет тотала
                cubature_sum += result['cubature']

                counter += 1

            # Тотал
            ws1[f'J{counter}'] = "Итого"
            ws1[f'J{counter}'].font = Font(bold=True)
            ws1[f'K{counter}'] = cubature_sum
            ws1[f'K{counter}'].font = Font(bold=True)

            self.set_border(ws1, f'A1:M{counter - 1}')
            self.set_print_settings(ws1)
            await session.close()
            return self.to_bytes(wb)

    async def get_dependent_detail_by_deleted_or_adjusted(self, from_date, to_date) -> str:
        async with db_helper.get_db_session() as session:
            # materials_list, client_companies_list, objects_list = [], [], []
            #
            # if materials:
            #     try:
            #         materials_list = [int(id) for id in materials.split(",")]
            #     except:
            #         raise BadRequestException("material are incorrect")
            # if client_companies:
            #     try:
            #         client_companies_list = [int(id) for id in client_companies.split(",")]
            #     except:
            #         raise BadRequestException("client companies are incorrect")
            # if objects:
            #     try:
            #         objects_list = [int(id) for id in objects.split(",")]
            #     except:
            #         raise BadRequestException("objects are incorrect")

            stmt = (
                select(
                    Weighing.first_at,
                    Weighing.second_at,
                    Company.company_type,
                    Company.name.label('company_name'),
                    Object.name.label('object_name'),
                    Transport.plate_number,
                    Material.name.label('material_name'),
                    Carrier.name.label('carrier_name'),
                    Weighing.clean_weight,
                    Weighing.brutto_weight,
                    Weighing.old_brutto_weight,
                    Weighing.tare_weight,
                    Weighing.old_tare_weight,
                    Weighing.deactivate_note,
                    Weighing.adjust_note,
                    Weighing.cubature,
                    User.fullname
                ).
                join(
                    Detail, Weighing.detail_id == Detail.id
                ).
                join(
                    Company, Detail.client_company_id == Company.id
                ).
                join(
                    Object, Detail.object_id == Object.id
                ).
                join(
                    Transport, Weighing.transport_id == Transport.id
                ).
                outerjoin(
                    Carrier, Transport.carrier_id == Carrier.id
                ).
                join(
                    Material, Detail.material_id == Material.id
                ).
                outerjoin(
                    User, Weighing.second_operator_id == User.id
                ).
                filter(
                    # Weighing.is_finished == True,
                    Weighing.is_depend == True,
                    # Weighing.clean_weight.isnot(None),
                    Weighing.second_at.between(from_date, to_date)
                ).
                order_by(
                    Weighing.second_at
                )
            )

            deactivated_stmt = stmt.filter(Weighing.is_active == False)
            raw = await session.execute(deactivated_stmt)
            deactivated_results = raw.mappings().all()

            adjusted_stmt = stmt.filter(Weighing.is_adjusted == True)
            raw = await session.execute(adjusted_stmt)
            adjusted_results = raw.mappings().all()

            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Data"

            width = 100
            height = 100
            ws1.column_dimensions['A'].width = 5
            ws1.row_dimensions[1].height = 50

            # Дата и время заезда/выезда
            ws1.column_dimensions['B'].width = width / 6

            # Заказчик
            ws1.column_dimensions['C'].width = width / 6

            # Объект
            ws1.column_dimensions['D'].width = width / 8

            # Номер машины
            ws1.column_dimensions['E'].width = width / 8.5

            # Груз
            ws1.column_dimensions['F'].width = width / 8

            # Перевозчик
            ws1.column_dimensions['G'].width = 11.7

            # Нетто
            ws1.column_dimensions['H'].width = width / 12

            # Брутто
            ws1.column_dimensions['I'].width = width / 13

            # Тара
            ws1.column_dimensions['J'].width = width / 13

            # Кубатура
            ws1.column_dimensions['K'].width = width / 11

            # Плотность
            ws1.column_dimensions['L'].width = 10.5

            # Диспетчер
            ws1.column_dimensions['M'].width = width / 8

            # Примечание
            ws1.column_dimensions['N'].width = width / 6

            from_date_title = from_date.strftime("%d-%m-%Y %H:%M")
            to_date_title = to_date.strftime("%d-%m-%Y %H:%M")

            ws1.merge_cells('A1:N1')
            ws1['A1'] = (f"Отчет по удаленным зависимым отвесам ТОО\n"
                         f"'BARAKAT INVEST' с {from_date_title} по {to_date_title}")
            ws1['A1'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

            ws1['A2'] = f"№"
            ws1['A2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['B2'] = (f"Дата и время\n"
                         f"заезда/выезда")
            ws1['B2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['C2'] = f"Заказчик"
            ws1['C2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['D2'] = f"Объект"
            ws1['D2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['E2'] = (f"Номер\n"
                         f"машины")
            ws1['E2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['F2'] = f"Марка б/р"
            ws1['F2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['G2'] = f"Перевозчик"
            ws1['G2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['H2'] = f"Нетто"
            ws1['H2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['I2'] = f"Брутто"
            ws1['I2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['J2'] = f"Тара"
            ws1['J2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['K2'] = f"Кубатура"
            ws1['K2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['L2'] = f"Плотность"
            ws1['L2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['M2'] = f"Диспетчер"
            ws1['M2'].alignment = Alignment(horizontal="center", vertical="center")

            ws1['N2'] = f"Примечание"
            ws1['N2'].alignment = Alignment(horizontal="center", vertical="center")

            counter = 3
            cubature_sum = 0
            last_cell: str

            for result in deactivated_results:
                from_date_data = result['first_at'].strftime("%d-%m-%Y %H:%M")
                to_date_data = result['second_at'].strftime("%d-%m-%Y %H:%M")

                # Счетчик
                ws1[f'A{counter}'] = counter - 2

                # Дата от и до
                ws1[f'B{counter}'] = (f"{from_date_data}\n"
                                      f"{to_date_data}")

                # Заказчик
                ws1[f'C{counter}'] = f"{result['company_type'].value} {result['company_name']}"

                # Объект
                ws1[f'D{counter}'] = result['object_name']

                # Номер машины
                ws1[f'E{counter}'] = result['plate_number']

                # Груз
                ws1[f'F{counter}'] = result['material_name']

                # Перевозчик
                ws1[f'G{counter}'] = result['carrier_name']

                # Нетто
                ws1[f'H{counter}'] = result['clean_weight']
                # ws1[f'H{counter}'].font = Font(bold=False)

                # Брутто
                ws1[f'I{counter}'] = result['brutto_weight']

                # Тара
                ws1[f'J{counter}'] = result['tare_weight']

                # Кубатура
                ws1[f'K{counter}'] = result['cubature']

                # Плотность
                if result['clean_weight']:
                    ws1[f'L{counter}'] = ceil(result['clean_weight'] / result['cubature'])
                else:
                    ws1[f'L{counter}'] = 0

                # Оператор
                ws1[f'M{counter}'] = result['fullname']

                # Подсчет тотала
                cubature_sum += result['cubature']

                # Примечание
                ws1[f'N{counter}'] = result['deactivate_note']

                counter += 1

            # Тотал
            ws1[f'J{counter}'] = "Итого"
            ws1[f'J{counter}'].font = Font(bold=True)
            ws1[f'K{counter}'] = cubature_sum
            ws1[f'K{counter}'].font = Font(bold=True)

            cubature_sum = 0
            self.set_border(ws1, f'A1:N{counter - 1}')
            counter += 2
            start_row = counter

            ws1.merge_cells(f'A{counter}:P{counter}')
            ws1[f'A{counter}'] = (f"Отчет по корректированным зависимым отвесам ТОО\n"
                                  f"'BARAKAT INVEST' с {from_date_title} по {to_date_title}")
            ws1.row_dimensions[counter].height = height / 2

            ws1[f'fA{counter}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            counter += 1
            ws1[f'A{counter}'] = f"№"
            ws1[f'A{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'B{counter}'] = (f"Дата и время\n"
                                  f"заезда/выезда")
            ws1[f'B{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'C{counter}'] = f"Заказчик"
            ws1[f'C{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'D{counter}'] = f"Объект"
            ws1[f'D{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'E{counter}'] = (f"Номер\n"
                                  f"машины")
            ws1[f'E{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'F{counter}'] = f"Марка б/р"
            ws1[f'F{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'G{counter}'] = f"Перевозчик"
            ws1[f'G{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'H{counter}'] = f"Нетто"
            ws1[f'H{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'I{counter}'] = f"Брутто"
            ws1[f'I{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'J{counter}'] = f"Тара"
            ws1[f'J{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'K{counter}'] = f"Кубатура"
            ws1[f'K{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'L{counter}'] = f"Плотность"
            ws1[f'L{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'M{counter}'] = f"Диспетчер"
            ws1[f'M{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'N{counter}'] = f"Примечание"
            ws1[f'N{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'O{counter}'] = f"Старое брутто"
            ws1[f'O{counter}'].alignment = Alignment(horizontal="center", vertical="center")

            ws1[f'P{counter}'] = f"Старая тара"
            ws1[f'P{counter}'].alignment = Alignment(horizontal="center", vertical="center")
            counter += 1

            for result in adjusted_results:
                from_date_data = result['first_at'].strftime("%d-%m-%Y %H:%M")
                to_date_data = result['second_at'].strftime("%d-%m-%Y %H:%M")

                # Счетчик
                ws1[f'A{counter}'] = counter - 6

                # Дата от и до
                ws1[f'B{counter}'] = (f"{from_date_data}\n"
                                      f"{to_date_data}")

                # Заказчик
                ws1[f'C{counter}'] = f"{result['company_type'].value} {result['company_name']}"

                # Объект
                ws1[f'D{counter}'] = result['object_name']

                # Номер машины
                ws1[f'E{counter}'] = result['plate_number']

                # Груз
                ws1[f'F{counter}'] = result['material_name']

                # Перевозчик
                ws1[f'G{counter}'] = result['carrier_name']

                # Нетто
                ws1[f'H{counter}'] = result['clean_weight']
                # ws1[f'H{counter}'].font = Font(bold=False)

                # Брутто
                ws1[f'I{counter}'] = result['brutto_weight']

                # Тара
                ws1[f'J{counter}'] = result['tare_weight']

                # Кубатура
                ws1[f'K{counter}'] = result['cubature']

                # Плотность
                ws1[f'L{counter}'] = ceil(result['clean_weight'] / result['cubature'])

                # Оператор
                ws1[f'M{counter}'] = result['fullname']

                # Подсчет тотала
                cubature_sum += result['cubature']

                # Примечание
                ws1[f'N{counter}'] = result['adjust_note']

                # Старое брутто
                ws1[f'O{counter}'] = result['old_brutto_weight']

                # Старое тара
                ws1[f'P{counter}'] = result['old_brutto_weight']

                counter += 1

            # Тотал
            ws1[f'J{counter}'] = "Итого"
            ws1[f'J{counter}'].font = Font(bold=True)
            ws1[f'K{counter}'] = cubature_sum
            ws1[f'K{counter}'].font = Font(bold=True)

            self.set_border(ws1, f'A{start_row}:P{counter - 1}')
            self.set_print_settings(ws1)
            await session.close()
            return self.to_bytes(wb)

    async def get_dependent_by_materials(self, from_date, to_date, report_type) -> str:
        async with db_helper.get_db_session() as session:

            try:
                stmt = select(ConcreteMixingPlant.name, ConcreteMixingPlant.ip_address)
                raw = await session.execute(stmt)
                data_from_util = raw.mappings().all()
                data_from_util = sorted(data_from_util, key=lambda x: int(x['name']))

                date_data = {
                    "from_date": str(from_date),
                    "to_date": str(to_date)
                }
                cmp_datas = {}

                for cmp in data_from_util:
                    try:
                        cmp_data = await get_data_for_materials_report(cmp['ip_address'], date_data)
                    except ClientConnectionError:
                        continue
                    cmp_datas[f'{cmp["name"]}'] = cmp_data
                # print(cmp_datas)
            except ClientConnectionError:
                pass

            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Data"
            yellow = PatternFill("solid", fgColor="FFFF00")
            green = PatternFill("solid", fgColor="00B050")
            red = PatternFill("solid", fgColor="FF0000")
            cyan = PatternFill("solid", fgColor="00B0F0")
            bold = Font(bold=True)

            start_row = 7
            start_col = 1

            # get material types in key and materials in value
            material_data = {}
            stmt = (
                select(
                    MaterialType.id,
                    MaterialType.name,
                )
                .filter(
                    MaterialType.is_for_dependent == True
                )
            )
            raw = await session.execute(stmt)
            results = raw.mappings().all()
            for result in results:
                stmt = select(
                    Material.name
                ).filter(
                    Material.material_type_id == result.id
                )
                raw = await session.execute(stmt)
                results_material = raw.scalars().all()
                material_data[result.name] = results_material

            # calculate last row
            num_keys = len(material_data)
            num_values = sum(len(values) for values in material_data.values())
            end_row = (num_keys + num_values) + start_row

            if report_type == "day":
                # calculate last col
                delta = to_date - from_date  # returns timedelta
                periods = []
                for i in range(delta.days + 1):
                    day = (from_date + timedelta(days=i)).strftime("%d-%m")
                    periods.append(day)
                # end_col = len(periods) + start_col + 1
                # every_period_data = len(periods) + start_col
                stmt = (
                    select(
                        Material.name.label('material_name'),
                        # MaterialType.name.label('material_type_name'),
                        func.DATE(Weighing.second_at).label('day'),
                        func.sum(Weighing.cubature).label('cubature_by_material'),
                    ).
                    join(
                        Detail, Weighing.detail_id == Detail.id
                    ).
                    join(
                        Material, Detail.material_id == Material.id
                    ).
                    join(
                        MaterialType, Material.material_type_id == MaterialType.id
                    ).
                    filter(
                        Weighing.is_active == True,
                        Weighing.is_finished == True,
                        Weighing.is_depend == True,
                        MaterialType.is_for_dependent == True,
                        Weighing.clean_weight.isnot(None),
                        Weighing.second_at.between(from_date, to_date),
                    ).
                    order_by(
                        'day'
                    ).
                    group_by(
                        'day', Material.name, MaterialType.name
                    )
                )
                raw = await session.execute(stmt)
                results = raw.mappings().all()
                new_data = []
                for result in results:
                    d = {}
                    d['material_name'] = result['material_name']
                    d['second_at'] = result['day'].strftime("%d-%m")
                    d['cubature_by_material'] = result['cubature_by_material']
                    new_data.append(d)

            if report_type == "month":
                # calculate last col
                periods = []
                # Проходим через каждый месяц
                cnt_from_date = from_date
                while cnt_from_date <= to_date:
                    month = cnt_from_date.strftime("%m-%Y")  # Форматируем месяц и год
                    periods.append(month)
                    # Увеличиваем дату на 1 месяц
                    if cnt_from_date.month == 12:
                        cnt_from_date = datetime.datetime(cnt_from_date.year + 1, 1, 1)
                    else:
                        cnt_from_date = datetime.datetime(cnt_from_date.year, cnt_from_date.month + 1, 1)

                stmt = (
                    select(
                        Material.name.label('material_name'),
                        func.concat(
                            extract('month', Weighing.second_at),
                            '-',
                            extract('year', Weighing.second_at),
                        ).label('date'),
                        func.sum(Weighing.cubature).label('cubature_by_material'),
                    ).
                    join(
                        Detail, Weighing.detail_id == Detail.id
                    ).
                    join(
                        Material, Detail.material_id == Material.id
                    ).
                    join(
                        MaterialType, Material.material_type_id == MaterialType.id
                    ).
                    filter(
                        Weighing.is_active == True,
                        Weighing.is_finished == True,
                        Weighing.is_depend == True,
                        MaterialType.is_for_dependent == True,
                        Weighing.clean_weight.isnot(None),
                        Weighing.second_at.between(from_date, to_date),
                    ).
                    group_by(
                        'date',
                        Material.name
                    )
                )
                raw = await session.execute(stmt)
                results = raw.mappings().all()
                new_data = []
                for result in results:
                    d = {}
                    d['material_name'] = result['material_name']
                    d['second_at'] = result['date']
                    d['cubature_by_material'] = result['cubature_by_material']
                    new_data.append(d)

            every_period_data = len(periods) + start_col
            end_col = len(periods) + start_col + 1

            cnt = 0
            in_material_type = True
            current_material_list = []
            material_data_iterator = iter(material_data)

            for row in ws1.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
                for cell in row:
                    if isinstance(cell, Cell):
                        if cell.column_letter == 'A':
                            if current_material_list:
                                cell.value = current_material_list.pop()
                                in_material_type = False

                                cell.fill = yellow
                                cell.font = bold
                            elif not current_material_list:
                                in_material_type = True
                            if in_material_type and cell.row < end_row - 1:
                                current_key = next(material_data_iterator)
                                # print(f"{current_key=}")
                                current_material_list = material_data[current_key][:]
                                cell.value = f"Марка {current_key}"
                                in_material_type = False

                                if cell.row != start_row:
                                    ws1.merge_cells(f'B{cell.row}:{get_column_letter(end_col - 1)}{cell.row}')
                                    ws1[f'B{cell.row}'].fill = yellow

                                cell.fill = yellow
                                cell.font = bold
                        elif cell.row == start_row and cell.column <= every_period_data:
                            cell.value = periods[cnt]
                            cnt += 1

                            cell.fill = yellow
                            cell.font = bold
                        if cell.column_letter == 'A' and cell.row == end_row or cell.column_letter == get_column_letter(
                                end_col) and cell.row == start_row:
                            cell.value = "Итого"
                        for item in new_data:
                            material_name = ws1[f'{get_column_letter(start_col)}{cell.row}'].value
                            second_at = ws1[f'{cell.column_letter}{start_row}'].value
                            if item['material_name'] == material_name and item['second_at'] in second_at:
                                cell.value = item['cubature_by_material']
                        if cell.column == end_col:
                            if cell.row != start_row:
                                cell.value = f"=SUM(B{cell.row}:{get_column_letter(cell.column - 1)}{cell.row})"
                            cell.fill = green
                        elif cell.row == end_row:
                            if cell.column != start_col:
                                cell.value = f"=SUM({cell.column_letter}{start_row + 1}:{cell.column_letter}{end_row - 1})"
                            cell.fill = green
            self.set_border(ws1, f'{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}')

            ws1.column_dimensions[get_column_letter(start_col)].width = 60
            # bsu
            start_col = end_col + 1
            ws1.column_dimensions[get_column_letter(start_col)].width = 60
            end_col = end_col + 9
            in_material_type = True
            current_material_list = []
            material_data_iterator = iter(material_data)

            office_col = get_column_letter(start_col + 1)
            ws1[f'{office_col}{start_row}'] = "Офис"
            ws1[f'{office_col}{start_row}'].fill = yellow

            chast_col = get_column_letter(start_col + 2)
            ws1[f'{chast_col}{start_row}'] = "Част."
            ws1.column_dimensions[chast_col].width = 15

            ws1[f'{chast_col}{start_row}'].fill = yellow

            general_col = get_column_letter(start_col + 3)
            ws1[f'{general_col}{start_row}'] = "Общее"
            ws1[f'{general_col}{start_row}'].fill = yellow

            difference_col = get_column_letter(start_col + 4)
            ws1[f'{difference_col}{start_row}'] = "Разница"
            ws1[f'{difference_col}{start_row}'].fill = red

            general_bsu_col = get_column_letter(start_col + 5)
            ws1[f'{general_bsu_col}{start_row}'] = "БСУ (общее)"
            ws1[f'{general_bsu_col}{start_row}'].fill = yellow

            bsu_start_col = start_col + 6
            bsu_end_col = bsu_start_col - 1
            # print(f"{data_from_util=}")
            for cmp in data_from_util:
                bsu_end_col += 1
                col = get_column_letter(bsu_end_col)
                ws1[f'{col}{start_row}'] = cmp.name
                ws1[f'{col}{start_row}'].fill = yellow

            for row in ws1.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
                for cell in row:
                    if isinstance(cell, Cell):
                        if cell.column_letter == get_column_letter(start_col):
                            if current_material_list:
                                cell.value = current_material_list.pop()
                                in_material_type = False

                                cell.fill = yellow
                                cell.font = bold
                            elif not current_material_list:
                                in_material_type = True
                            if in_material_type and cell.row < end_row - 1:
                                if cell.row != start_row:
                                    ws1.merge_cells(
                                        f'{get_column_letter(start_col - 1)}{cell.row}:{get_column_letter(start_col)}{cell.row}')
                                    # ws1[f'{get_column_letter(start_col - 1)}{cell.row}'].fill = yellow

                                current_key = next(material_data_iterator)
                                # print(f"{current_key=}")
                                current_material_list = material_data[current_key]

                                if cell.row == start_row:
                                    cell.value = f"Марка {current_key}"
                                else:
                                    ws1[f'{get_column_letter(cell.column - 1)}{cell.row}'] = f"Марка {current_key}"
                                    ws1.merge_cells(
                                        f'{get_column_letter(cell.column + 1)}{cell.row}:{get_column_letter(end_col)}{cell.row}')
                                    ws1[f'{get_column_letter(start_col - 1)}{cell.row}'].fill = yellow

                                in_material_type = False

                                cell.fill = yellow
                                cell.font = bold
                        if cell.row != start_row:
                            if cell.column_letter == office_col:
                                if not isinstance(ws1[f'{get_column_letter(cell.column + 1)}{cell.row}'], MergedCell):
                                    cell.fill = green
                                    cell.value = ws1[f'{get_column_letter(cell.column - 2)}{cell.row}'].value
                                else:
                                    cell.fill = yellow
                            elif cell.column_letter == chast_col:
                                cell.fill = green
                                if cell.row == end_row:
                                    cell.value = f"=SUM({cell.column_letter}{start_row + 1}:{cell.column_letter}{cell.row - 1})"
                            elif cell.column_letter == general_col:
                                cell.value = f"={office_col}{cell.row}+{chast_col}{cell.row}"
                                cell.fill = green
                            elif cell.column_letter == difference_col:
                                cell.value = f"={general_col}{cell.row}-{general_bsu_col}{cell.row}"
                                if cell.row == end_row:
                                    cell.fill = red
                                    cell.value = f"=SUM({cell.column_letter}{start_row + 1}:{cell.column_letter}{cell.row - 1})"
                            elif cell.column_letter == general_bsu_col:
                                cell.value = f"=SUM({get_column_letter(bsu_start_col)}{cell.row}:{get_column_letter(bsu_end_col)}{cell.row})"
                                if cell.row == end_row:
                                    cell.fill = green
                                    cell.value = f"=SUM({cell.column_letter}{start_row + 1}:{cell.column_letter}{cell.row - 1})"
                            elif bsu_start_col <= cell.column <= end_col:
                                material_name = ws1[f'{get_column_letter(start_col)}{cell.row}'].value

                                cmp_name = ws1[f'{cell.column_letter}{start_row}'].value
                                # print(cmp_name)
                                # print(cmp_datas)
                                try:
                                    cmp = cmp_datas[cmp_name]
                                    if not cmp:
                                        cell.fill = red
                                        continue
                                except KeyError:
                                    # print(f"{cmp_name} is incorrect")
                                    cell.fill = red
                                    continue
                                for cmp_data in cmp:
                                    if cmp_data['material_name'] and cmp_data['material_name'] == material_name:
                                        cell.value = cmp_data['sum']
                                if cell.row == end_row:
                                    cell.fill = green
                                    cell.value = f"=SUM({cell.column_letter}{start_row + 1}:{cell.column_letter}{cell.row - 1})"

            self.set_border(ws1, f'{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}')

            ws1[f'{get_column_letter(start_col)}{end_row + 2}'] = "План"
            ws1[f'{get_column_letter(start_col)}{end_row + 3}'] = ""
            ws1[f'{get_column_letter(start_col)}{end_row + 2}'].fill = cyan

            ws1[f'{get_column_letter(start_col + 1)}{end_row + 2}'] = "Факт"
            ws1[f'{get_column_letter(start_col + 1)}{end_row + 2}'].fill = cyan
            ws1[f'{get_column_letter(start_col + 1)}{end_row + 3}'] = ws1[
                f'{get_column_letter(start_col - 1)}{end_row}'].value
            ws1[f'{get_column_letter(start_col + 1)}{end_row + 3}'].fill = yellow

            ws1[f'{get_column_letter(start_col + 2)}{end_row + 2}'] = "Отставание"
            ws1[f'{get_column_letter(start_col + 2)}{end_row + 2}'].fill = cyan
            ws1[
                f'{get_column_letter(start_col + 2)}{end_row + 3}'] = f"={get_column_letter(start_col - 1)}{end_row}-{get_column_letter(start_col)}{end_row + 3}"
            ws1[f'{get_column_letter(start_col + 2)}{end_row + 3}'].fill = red

            self.set_border(ws1,
                            f'{get_column_letter(start_col)}{end_row + 2}:{get_column_letter(start_col + 2)}{end_row + 3}')

            middle_col = ceil(end_col / 2) - 6
            ws1.merge_cells(
                f'{get_column_letter(middle_col)}5:{get_column_letter(middle_col + 6)}5')
            from_date_title = from_date.strftime("%d-%m-%Y %H:%M")
            to_date_title = to_date.strftime("%d-%m-%Y %H:%M")
            ws1[
                f'{get_column_letter(middle_col)}5'] = f"Реализация и производство от {from_date_title} до {to_date_title}"

            right_alignment = Alignment(horizontal='right', vertical='center')

            # ws1.merge_cells(
            #     f'{general_bsu_col}1:{bsu_3_col}1')
            # ws1[f'{general_bsu_col}1'] = 'Директор ТОО "Баракат Инвест"'
            # ws1[f'{general_bsu_col}1'].alignment = right_alignment
            #
            # ws1.merge_cells(
            #     f'{general_bsu_col}3:{bsu_3_col}3')
            # ws1[f'{general_bsu_col}3'] = '_______________ Гапбасов Ж.А.'
            # ws1[f'{general_bsu_col}3'].alignment = right_alignment

            self.set_bold(ws1, f'A1:{get_column_letter(end_col)}{end_row + 3}')
            self.set_print_settings(ws1)
            await session.close()
            return self.to_bytes(wb)

    async def generate_weighing_act_xlsx(self, weighing_id: int) -> str:
        """
        Генерирует XLSX акт взвешивания по id отвеса и возвращает путь к файлу
        """
        async with db_helper.get_db_session() as session:
            from models.weighing import Weighing
            from models.detail import Detail
            from models.company import Company
            from models.material import Material
            from models.transport import Transport
            from models.driver import Driver
            from models.carrier import Carrier
            from models.user import User

            weighing = await session.get(Weighing, weighing_id)
            if not weighing:
                raise BadRequestException("Отвес не найден")
            detail = await session.get(Detail, weighing.detail_id)
            seller = await session.get(Company, detail.seller_company_id)
            client = await session.get(Company, detail.client_company_id)
            material = await session.get(Material, detail.material_id)
            transport = await session.get(Transport, weighing.transport_id)
            driver = await session.get(Driver, weighing.driver_id) if weighing.driver_id else None
            carrier = await session.get(Carrier, transport.carrier_id) if transport and transport.carrier_id else None
            first_operator = await session.get(User, weighing.first_operator_id) if weighing.first_operator_id else None
            second_operator = await session.get(User, weighing.second_operator_id) if weighing.second_operator_id else None

            wb = Workbook()
            ws = wb.active
            ws.title = "Акт взвешивания"

            ws["A1"] = f"Акт взвешивания №{weighing.id}"
            ws["A2"] = f"Поставщик: {getattr(seller, 'name', '')}"
            ws["A3"] = f"Заказчик: {getattr(client, 'name', '')}"
            ws["A4"] = f"Гос.номер: {getattr(transport, 'plate_number', '')}"
            ws["A5"] = f"Водитель: {getattr(driver, 'fullname', '')}"
            ws["A6"] = f"Материал: {getattr(material, 'name', '')}"
            ws["A7"] = f"Перевозчик: {getattr(carrier, 'name', '')}"

            ws["A9"] = "Данные по весу:"
            ws.append(["Брутто, кг", "Тара, кг", "Нетто, кг", "Номер силоса", "Кол-во мешков"])
            ws.append([
                weighing.brutto_weight or '-',
                weighing.tare_weight or '-',
                weighing.netto_weight or '-',
                weighing.silo_number or '-',
                weighing.bag_details or '-',
            ])

            ws.append([])
            ws.append(["", "Первое взвешивание", "Второе взвешивание"])
            ws.append([
                "Оператор",
                getattr(first_operator, 'fullname', '-') if first_operator else '-',
                getattr(second_operator, 'fullname', '-') if second_operator else '-',
            ])
            ws.append([
                "Дата и время",
                weighing.first_at.strftime("%d-%m-%Y %H:%M") if weighing.first_at else '-',
                weighing.second_at.strftime("%d-%m-%Y %H:%M") if weighing.second_at else '-',
            ])

            file_path = f"media/act_{weighing.id}.xlsx"
            wb.save(file_path)
            return file_path
